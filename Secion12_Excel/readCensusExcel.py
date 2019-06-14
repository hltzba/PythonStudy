# encoding:utf-8
#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for each county.
import openpyxl,pprint,os
print('Opening workbook...')
print(os.getcwd())
wb=openpyxl.load_workbook('censuspopdata.xlsx')
sheet=wb['Population by Census Tract']
countydata={}
# TODO:Fill in countydata with each county's population and tracts.
print('Reading rows...')
print(sheet.max_row)

for row in range(2,sheet.max_row+1):
    # Each row in spreadsheet has data for one census tract.
    State=sheet['B'+str(row)].value
    County=sheet['C'+str(row)].value
    Pop=sheet['D'+str(row)].value
    #Make sure the key for this state exists.
    countydata.setdefault(State,{})
    #Make sure the key for this county in this state exists.
    countydata[State].setdefault(County,{'tracts':0,'pop':0})
    #Each row represents one census tract,so increment by one.
    countydata[State][County]['tracts']+=1
    #Increase the county pop by the pop in this census tract.
    countydata[State][County]['pop']+=int(Pop)
# TODO:Open a new text file and write the contents of countydata to it.
print('Writing results...')
resultFile=open('census2010.py','w')
resultFile.write('allData = '+pprint.pformat(countydata))
resultFile.close()
print('Done.')