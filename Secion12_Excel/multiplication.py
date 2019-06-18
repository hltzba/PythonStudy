# encoding:utf-8
#! python3
# multiplicaiton.py - 99乘法口诀表
import openpyxl
from openpyxl.styles import Font
wb=openpyxl.Workbook()
ws=wb['Sheet']
for row in range (1,11):
    for col in range(1,11):
        colchar=openpyxl.utils.get_column_letter(col)
        positionName=colchar+str(row)
        if positionName=='A1':
            continue
        if row==1:
           ws[positionName].value=col-1
           ws[positionName].font=Font(bold=True)
        elif col==1:
            ws[positionName].value=row-1
            ws[positionName].font=Font(bold=True)
        else:
            pos1Name=colchar+'1'
            value1=int(ws[pos1Name].value)
            pos2Name='A'+str(row)
            value2=int(ws[pos2Name].value)
            ws[positionName].value=value1*value2
wb.save('99口诀表.xlsx')

