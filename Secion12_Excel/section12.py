# encoding:utf-8
#! python3
# section12.py
import openpyxl
from openpyxl.styles import Font
def readbaseinfo():
    wb=openpyxl.load_workbook('E:\\青洲大桥交通量统计报表(对比).xlsx')
    print(wb.sheetnames)
    sheet1=wb['Sheet1']
    print('Sheet1 Title:'+sheet1.title)
    #print(sheet1['A1'].value)
    print(sheet1.cell(2,1).value)    
    print('Sheet1 highest row:%s,highest cloumn:%s.'%(sheet1.max_row,sheet1.max_column))

def indexParse():
    print(openpyxl.utils.get_column_letter(1))
    print(openpyxl.utils.column_index_from_string('A'))

def sliceData():
    wb=openpyxl.load_workbook('E:\\青洲大桥交通量统计报表(对比).xlsx')
    sheet1=wb['Sheet1']
    data=sheet1['B6':'F10']
    for rowOfCellObjs in data:
        for cellobj in rowOfCellObjs:
            print(cellobj.coordinate,cellobj.value)
        print('ROW END'.center(12,'-'))

def createWorkbook():
    wb=openpyxl.Workbook()
    sheet=wb.get_active_sheet()
    sheet.title='openpyxlDemo'
    print(wb.sheetnames)
    sheet['A1']=1
    sheet['A2']='1'
    wb.save('demo.xlsx')

def createSheetByFuncParam():
    wb=openpyxl.Workbook()
    wb.create_sheet(index=0,title='1st Sheet')
    wb.create_sheet(index=1,title='2rd Sheet')
    wb.create_sheet(index=2,title='temp Sheet')
    print(wb.sheetnames)
    wb.remove(wb['temp Sheet'])
    print(wb.sheetnames)
    wb['1st Sheet']['A1']='Hello world!'
    print(wb['1st Sheet']['A1'].value)

def cellFontStyleDemo():
    wb=openpyxl.Workbook()
    sheet=wb['Sheet']
    italic24Font=Font(size=24,italic=True)
    #sheet.column_dimensions['A'].font=Font(size=24,italic=True)
    col=sheet.column_dimensions['A']
    col.width=85
    sheet['A1'].font=italic24Font
    sheet.merge_cells('A1:C1')
    sheet['A1']='Hello world!'
    
    wb.save('styles.xlsx')

def chartsDemo():
    wb=openpyxl.Workbook()
    sheet=wb['Sheet']
    for i in range(1,11):
        sheet['A'+str(i)]=i
    refObj=openpyxl.chart.Reference(sheet,min_row=1,min_col=1,max_row=10,max_col=1)
    seriesObj=openpyxl.chart.Series(refObj,title='First Series')
    chartObj=openpyxl.chart.BarChart()
    chartObj.append(seriesObj)
    sheet.add_chart(chartObj,'B1')
    wb.save('chartdemo.xlsx')


#readbaseinfo()
indexParse()
#sliceData()
#createWorkbook()
#createSheetByFuncParam()
#cellFontStyleDemo()
#chartsDemo()